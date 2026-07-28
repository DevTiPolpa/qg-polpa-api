"""
export_excel.py - Gera a planilha final: 3 abas por classe (Livre/Revisar/Bloqueada)
+ 1 aba de Resumo com a avaliacao da lista.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SATURATION_THRESHOLD = 0.65

_FILL = {
    "LIVRE": PatternFill("solid", fgColor="C6EFCE"),
    "REVISAR": PatternFill("solid", fgColor="FFEB9C"),
    "BLOQUEADA": PatternFill("solid", fgColor="FFC7CE"),
}
_FONT_HEADER = Font(bold=True)

_COLUNAS_BASE = ["nome", "marca", "cidade", "uf", "cnpj", "site", "telefone", "email"]
_COLUNAS_EXTRA = {
    "REVISAR": ["motivo", "match_ref_nome", "match_score", "responsavel"],
    "BLOQUEADA": ["motivo", "fonte_bloqueio", "match_ref_nome", "responsavel"],
}


def _write_sheet(wb: Workbook, titulo: str, itens: list[dict], classe: str):
    ws = wb.create_sheet(titulo)
    colunas = _COLUNAS_BASE + _COLUNAS_EXTRA.get(classe, [])
    for col_idx, nome_col in enumerate(colunas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=nome_col)
        cell.font = _FONT_HEADER
        cell.fill = _FILL[classe]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(colunas))}1"

    for row_idx, item in enumerate(itens, start=2):
        for col_idx, nome_col in enumerate(colunas, start=1):
            ws.cell(row=row_idx, column=col_idx, value=item.get(nome_col))

    for col_idx, nome_col in enumerate(colunas, start=1):
        largura = max(14, min(50, len(nome_col) + 4))
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    return ws


def build_workbook(itens_classificados: list[dict], avaliacao: dict) -> Workbook:
    """itens_classificados: lista de dicts com nome/marca/cidade/uf/cnpj/site/telefone/
    email/classe/motivo/fonte_bloqueio/match_ref_nome/match_score/responsavel.
    avaliacao: dict com as chaves calculadas em evaluation.py."""
    wb = Workbook()
    wb.remove(wb.active)

    livres = [i for i in itens_classificados if i["classe"] == "LIVRE"]
    revisar = [i for i in itens_classificados if i["classe"] == "REVISAR"]
    bloqueadas = [i for i in itens_classificados if i["classe"] == "BLOQUEADA"]

    _write_sheet(wb, "Livre", livres, "LIVRE")
    _write_sheet(wb, "Revisar", revisar, "REVISAR")
    _write_sheet(wb, "Bloqueada", bloqueadas, "BLOQUEADA")

    ws = wb.create_sheet("Resumo", 0)
    linhas = [
        ("Total gerado pela Manus", avaliacao["total_manus"]),
        ("Livres para prospeccao", avaliacao["total_livre"]),
        ("Para revisar", avaliacao["total_revisar"]),
        ("Bloqueadas", avaliacao["total_bloqueada"]),
        ("  - bloqueadas: cliente ativo", avaliacao["bloqueada_cliente_ativo"]),
        ("  - bloqueadas: ja e lead no CRM", avaliacao["bloqueada_crm_lead"]),
        ("  - bloqueadas: ja e negocio no CRM", avaliacao["bloqueada_crm_deal"]),
        ("% aderencia ao criterio pedido", f"{avaliacao['pct_aderencia']:.0%}"),
        ("Taxa de aproveitamento liquido (livre/total)", f"{avaliacao['taxa_aproveitamento']:.0%}"),
        ("% com CNPJ preenchido", f"{avaliacao['pct_cnpj_presente']:.0%}"),
        ("% com contato preenchido (site/tel/email)", f"{avaliacao['pct_contato_presente']:.0%}"),
    ]
    for row_idx, (rotulo, valor) in enumerate(linhas, start=1):
        ws.cell(row=row_idx, column=1, value=rotulo).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=valor)
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 20

    if avaliacao["saturacao_alerta"]:
        row = len(linhas) + 2
        cell = ws.cell(
            row=row, column=1,
            value=(
                "ALERTA: mais de 65% da lista ja e bloqueada/revisar - "
                "este segmento/criterio pode estar saturado. Considere mudar de "
                "segmento ou aplicacao."
            ),
        )
        cell.font = Font(bold=True, color="9C0006")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)

    return wb
